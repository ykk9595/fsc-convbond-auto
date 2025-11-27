#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
fsc_convbond_all_in_one.py

整合流程：
  1) 依「今天日期」組出金管會「申報案件彙總表」Excel 下載網址：
       https://www.fsc.gov.tw/userfiles/file/YYYYMMDD申報案件彙總表.xlsx
  2) 下載並讀取 Excel：
       - 第 1 列：大標題（申報案件辦理情形彙總表(新聞稿)）
       - 第 2 列：欄位名稱（證券代號、公司型態、結案類型、公司名稱、案件類別、收文日期、生效日期…）
  3) 在「案件類別」欄位中，篩選包含「轉換公司債」的列
       - 只保留欄位：證券代號、公司型態、結案類型、公司名稱、收文日期、生效日期
       - 另存成 CSV：fsc_convbond_YYYYMMDD.csv
  4) 讀取 fsc_convbond_YYYYMMDD.csv，依欄位：
        A：證券代號
        E：收文日期
        F：生效日期
     透過 Yahoo Finance 取得：
        G：收文日期當天股價
        H：生效日期當天股價
        I：今日股價（最新收盤）
        J：收文日期成交張數
        K：生效日期成交張數
        L：今日成交張數（最新）
     並輸出：
        fsc_convbond_YYYYMMDD_with_price.xlsx
        fsc_convbond_YYYYMMDD_last20.xlsx
  5) 跑完後：用 LINE Bot 發一則摘要訊息給你
"""

import datetime as dt
import io
from typing import Optional
from pathlib import Path
import re
import os
import json

import pandas as pd
import requests
import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook


# ========= LINE Bot 設定 =========
# 在 GitHub Actions 的 Secrets 設兩個：
#   LINE_CHANNEL_ACCESS_TOKEN
#   LINE_USER_ID
LINE_CHANNEL_ACCESS_TOKEN = os.environ.get("LINE_CHANNEL_ACCESS_TOKEN", "")
LINE_USER_ID = os.environ.get("LINE_USER_ID", "")


def send_line_message(text: str):
    """用 LINE Messaging API push 一則文字訊息。"""
    if not LINE_CHANNEL_ACCESS_TOKEN or not LINE_USER_ID:
        print("[WARN] LINE 環境變數未設定，略過推播。")
        return

    url = "https://api.line.me/v2/bot/message/push"
    headers = {
        "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }

    # LINE 單則訊息上限 2000 字元，保守抓 1800
    if len(text) > 1800:
        text = text[:1800] + "\n...(訊息過長已截斷)"

    body = {
        "to": LINE_USER_ID,
        "messages": [
            {"type": "text", "text": text}
        ],
    }

    try:
        resp = requests.post(url, headers=headers, data=json.dumps(body), timeout=10)
        print(f"[LINE] status={resp.status_code}, body={resp.text}")
    except Exception as e:
        print(f"[LINE] 推播失敗：{e}")


# ========= 申報案件下載 + 篩選設定 =========

BASE_URL = "https://www.fsc.gov.tw/userfiles/file"

# 如果只想抓某一家公司的轉換債，可以在這裡填公司關鍵字
# 例如 TARGET_COMPANY = "凌航科技"
# 如果要抓全部轉換公司債案件，就設成 None
TARGET_COMPANY: Optional[str] = None  # 或改成 "凌航科技"


def build_daily_excel_url(date: dt.date) -> str:
    """依日期組出金管會每日申報案件彙總表的下載網址。"""
    date_str = date.strftime("%Y%m%d")
    filename = f"{date_str}申報案件彙總表.xlsx"
    return f"{BASE_URL}/{filename}"


def download_and_parse_excel(url: str) -> pd.DataFrame:
    """下載 Excel，並轉成欄位乾淨的 DataFrame。"""
    print(f"[INFO] 下載：{url}")
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()

    # header=1 表示用第 2 列當欄位（第一列是大標題）
    raw_df = pd.read_excel(io.BytesIO(resp.content), header=1)

    # raw_df 的第 0 列其實是「真正欄位名稱」（證券代號、公司型態…）
    # 所以再用第 0 列 rename 一次，並把那列丟掉
    df = raw_df.rename(columns=raw_df.iloc[0]).drop(0).reset_index(drop=True)

    print("[INFO] 欄位名稱：", list(df.columns))
    return df


def filter_conv_bond(df: pd.DataFrame, company_keyword: Optional[str] = None) -> pd.DataFrame:
    """從全部案件中篩選『案件類別包含：轉換公司債』，可選擇再篩公司名稱。"""
    if "案件類別" not in df.columns:
        raise KeyError("找不到『案件類別』欄位，請檢查今日 Excel 格式是否有變化。")

    # 先抓出轉換公司債相關案件（含「轉換公司債」三個字即可）
    mask_cb = df["案件類別"].astype(str).str.contains("轉換公司債", na=False)
    cb_df = df[mask_cb].copy()

    if cb_df.empty:
        print("[INFO] 今日申報案件中沒有『轉換公司債』相關案件。")
        return cb_df

    # 若有指定公司，再用公司名稱做第二層篩選
    if company_keyword:
        if "公司名稱" not in cb_df.columns:
            raise KeyError("找不到『公司名稱』欄位，請檢查今日 Excel 格式是否有變化。")

        mask_company = cb_df["公司名稱"].astype(str).str.contains(company_keyword, na=False)
        cb_df = cb_df[mask_company].copy()

        if cb_df.empty:
            print(f"[INFO] 今日雖有『轉換公司債』案件，但沒有公司名稱包含「{company_keyword}」。")

    return cb_df


def generate_convbond_csv_for_today() -> Optional[Path]:
    """
    執行「程式1」邏輯：
      - 下載今天申報案件彙總表
      - 篩選『轉換公司債』案件
      - 輸出 fsc_convbond_YYYYMMDD.csv
    回傳：CSV 的 Path，如沒有資料則回傳 None
    """
    today = dt.date.today()
    print(f"[INFO] 今日日期：{today}")

    url = build_daily_excel_url(today)

    try:
        df = download_and_parse_excel(url)
    except Exception as e:
        print(f"[ERROR] 無法下載或讀取 Excel：{e}")
        return None

    conv_df = filter_conv_bond(df, TARGET_COMPANY)

    if conv_df.empty:
        # 沒有任何轉換公司債案件
        return None

    # 只保留你要看的欄位
    wanted_cols = ["證券代號", "公司型態", "結案類型", "公司名稱", "收文日期", "生效日期"]

    # 確認欄位存在
    for col in wanted_cols:
        if col not in conv_df.columns:
            raise KeyError(f"找不到欄位：{col}，請檢查今日 Excel 是否有變更。")

    result = conv_df[wanted_cols].copy()

    print("\n========== 今日轉換公司債案件 ==========")
    print(result.to_string(index=False))
    print("=====================================\n")

    # 另存 CSV 檔給你留資料
    out_name = f"fsc_convbond_{today.strftime('%Y%m%d')}.csv"
    # utf-8-sig 方便用 Excel 開檔不會亂碼
    result.to_csv(out_name, index=False, encoding="utf-8-sig")
    print(f"[INFO] 已輸出 CSV：{out_name}")

    return Path(out_name)


# ========= 股價填入（原程式2） =========

# 欄位設定
COL_CODE = "A"
COL_RECV_DATE = "E"
COL_EFF_DATE = "F"
COL_RECV_PRICE = "G"
COL_EFF_PRICE = "H"
COL_TODAY_PRICE = "I"
COL_RECV_VOL = "J"
COL_EFF_VOL = "K"
COL_TODAY_VOL = "L"

HEADER_ROW = 1
START_ROW = 2


def parse_date(cell_value):
    """將民國年 / 西元年字串轉換成 datetime.date"""
    if cell_value is None or cell_value == "":
        return None

    if isinstance(cell_value, dt.datetime):
        return cell_value.date()
    if isinstance(cell_value, dt.date):
        return cell_value

    s = str(cell_value).strip()

    # 民國年：1141021
    if s.isdigit():
        if len(s) == 7:
            try:
                y = int(s[:3]) + 1911
                m = int(s[3:5])
                d = int(s[5:7])
                return dt.date(y, m, d)
            except:
                pass
        if len(s) == 8:
            try:
                return dt.datetime.strptime(s, "%Y%m%d").date()
            except:
                pass

    # 民國有符號
    roc = re.match(r"^(\d{3})[./-](\d{1,2})[./-](\d{1,2})$", s)
    if roc:
        try:
            y = int(roc.group(1)) + 1911
            m = int(roc.group(2))
            d = int(roc.group(3))
            return dt.date(y, m, d)
        except:
            pass

    # 常見西元格式
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except:
            continue

    return None


def get_yahoo_ohlcv_by_date(ticker, date):
    """抓指定日期 OHLCV（只有 1 根 K）"""
    if date is None:
        return None, None

    try:
        df = yf.download(
            ticker,
            start=date,
            end=date + dt.timedelta(days=1),
            progress=False,
            auto_adjust=False
        )
    except Exception as e:
        print(f"[WARN] 下載 {ticker} {date} 失敗：{e}")
        return None, None

    if df.empty:
        return None, None

    close = float(df["Close"].iloc[0])
    volume = float(df["Volume"].iloc[0])
    return close, volume


def get_yahoo_latest_ohlcv(ticker):
    """抓最近一個交易日 OHLCV（最新股價）"""
    try:
        df = yf.download(
            ticker,
            period="5d",
            interval="1d",
            progress=False,
            auto_adjust=False
        )
    except Exception as e:
        print(f"[WARN] 下載最新 {ticker} 失敗：{e}")
        return None, None

    if df.empty:
        return None, None

    close = float(df["Close"].iloc[-1])
    volume = float(df["Volume"].iloc[-1])
    return close, volume


def get_tw_ohlcv_by_date(code, date):
    """台股（先 TW 再 TWO）"""
    if date is None:
        return None, None

    code = str(code).strip()

    close, vol = get_yahoo_ohlcv_by_date(code + ".TW", date)
    if close is not None:
        return close, vol

    close, vol = get_yahoo_ohlcv_by_date(code + ".TWO", date)
    return close, vol


def get_tw_latest_ohlcv(code, cache):
    code = str(code).strip()

    if code in cache:
        return cache[code]

    close, vol = get_yahoo_latest_ohlcv(code + ".TW")
    if close is None:
        close, vol = get_yahoo_latest_ohlcv(code + ".TWO")

    cache[code] = (close, vol)
    return close, vol


def fill_prices_for_file(csv_path: Path):
    """
    執行「程式2」邏輯：
      - 讀取 fsc_convbond_YYYYMMDD.csv （或 .xlsx）
      - 依收文 / 生效 / 今日，補上股價 + 成交量
      - 輸出 *_with_price.xlsx 與 *_last20.xlsx
    回傳：(with_price_path, last20_path, n_rows)
    """
    original_path = csv_path

    # 允許 .csv 或 .xlsx
    if original_path.suffix.lower() == ".csv":
        df = pd.read_csv(original_path)
        xlsx_path = original_path.with_suffix(".xlsx")
        df.to_excel(xlsx_path, index=False)
    else:
        xlsx_path = original_path

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    today = dt.date.today()

    def normalize_date(d):
        if d is None:
            return None
        if d.year < 1990 or d > today:
            return None
        return d

    # 標題列設定
    def set_header(cell, text, bold=True, color=None, bg=None):
        cell.value = text
        cell.alignment = Alignment(horizontal="center")
        if bold:
            cell.font = Font(bold=True, color=color if color else "000000")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)

    set_header(ws[f"{COL_RECV_PRICE}1"], "收文日期當天股價", True, "FFFFFF", "0000FF")
    set_header(ws[f"{COL_EFF_PRICE}1"], "生效日期當天股價", True, "000000", "FFFF00")
    set_header(ws[f"{COL_TODAY_PRICE}1"], "今日股價", True, "FFFFFF", "000000")

    set_header(ws[f"{COL_RECV_VOL}1"], "收文日期成交張數")
    set_header(ws[f"{COL_EFF_VOL}1"], "生效日期成交張數")
    set_header(ws[f"{COL_TODAY_VOL}1"], "今日成交張數")

    # 最新資料 cache
    latest_cache = {}

    # ===== 逐列處理 =====
    row = START_ROW

    while True:
        code = ws[f"{COL_CODE}{row}"].value
        if code is None or str(code).strip() == "":
            break

        recv_raw = ws[f"{COL_RECV_DATE}{row}"].value
        eff_raw = ws[f"{COL_EFF_DATE}{row}"].value

        recv_date = normalize_date(parse_date(recv_raw))
        eff_date = normalize_date(parse_date(eff_raw))

        # 收文
        recv_close, recv_vol_sh = get_tw_ohlcv_by_date(code, recv_date)
        recv_lots = round(recv_vol_sh / 1000) if recv_vol_sh else None
        ws[f"{COL_RECV_PRICE}{row}"].value = recv_close
        ws[f"{COL_RECV_VOL}{row}"].value = recv_lots

        # 生效
        eff_close, eff_vol_sh = get_tw_ohlcv_by_date(code, eff_date)
        eff_lots = round(eff_vol_sh / 1000) if eff_vol_sh else None
        ws[f"{COL_EFF_PRICE}{row}"].value = eff_close
        ws[f"{COL_EFF_VOL}{row}"].value = eff_lots

        # 今日（最新）
        today_close, today_vol_sh = get_tw_latest_ohlcv(code, latest_cache)
        today_lots = round(today_vol_sh / 1000) if today_vol_sh else None
        ws[f"{COL_TODAY_PRICE}{row}"].value = today_close
        ws[f"{COL_TODAY_VOL}{row}"].value = today_lots

        row += 1

    # 實際資料列數（不含標題）
    n_rows = (row - START_ROW) if row > START_ROW else 0

    # ===== 存檔（避免檔案占用）=====
    base = xlsx_path.stem + "_with_price"
    out_path = xlsx_path.with_name(base + ".xlsx")

    idx = 1
    while True:
        try:
            wb.save(out_path)
            break
        except PermissionError:
            out_path = xlsx_path.with_name(f"{base}_{idx}.xlsx")
            idx += 1

    print(f"✔ 全資料已輸出：{out_path}")

    # ===== 產生最後 20 筆 =====
    wb2 = openpyxl.load_workbook(out_path)
    ws2 = wb2[wb2.sheetnames[0]]

    last_row = ws2.max_row
    start_last20 = max(START_ROW, last_row - 19)

    wb_last20 = Workbook()
    ws_last20 = wb_last20.active
    ws_last20.title = "last20"

    # 複製標題
    for col in range(1, ws2.max_column + 1):
        ws_last20.cell(row=1, column=col).value = ws2.cell(row=1, column=col).value

    # 複製最後 20 列
    to_row = 2
    for r in range(start_last20, last_row + 1):
        for c in range(1, ws2.max_column + 1):
            ws_last20.cell(row=to_row, column=c).value = ws2.cell(row=r, column=c).value
        to_row += 1

    last20_path = out_path.with_name(xlsx_path.stem + "_last20.xlsx")
    wb_last20.save(last20_path)

    print(f"✔ 最後 20 筆已另存：{last20_path}")

    return out_path, last20_path, n_rows


def build_summary_message(today: dt.date, csv_path: Path,
                          with_price_path: Path, last20_path: Path,
                          n_rows: int) -> str:
    """
    組成要推播到 LINE 的文字摘要：
      - 不再提到檔案名稱
      - 直接將「最後 20 筆」完整列在訊息裡
      - 資料來源：last20 的 Excel（含股價欄位）
    """
    try:
        # 直接讀取 last20 的 Excel（最多 20 筆）
        df = pd.read_excel(last20_path)
    except Exception as e:
        print(f"[WARN] 讀取 last20 Excel 失敗，無法組成明細：{e}")
        df = None

    lines = []

    if df is not None and not df.empty:
        # 逐筆列出（last20.xlsx 本身就只有最後 20 筆）
        for _, row in df.iterrows():
            code = row.get("證券代號", "")
            name = row.get("公司名稱", "")
            recv = row.get("收文日期", "")
            eff = row.get("生效日期", "")

            # 這三個欄位是在 fill_prices_for_file 裡自己加的標題
            recv_px = row.get("收文日期當天股價", "")
            eff_px = row.get("生效日期當天股價", "")
            today_px = row.get("今日股價", "")

            # 每一筆一行，你可以依喜好調整格式
            line = (
                f"{code} {name}\n"
                f"  收文:{recv}  生效:{eff}\n"
                f"  收文價:{recv_px}  生效價:{eff_px}  今日價:{today_px}"
            )
            lines.append(line)

    detail_str = "\n\n".join(lines) if lines else "（無明細或讀取失敗）"

    msg = (
        "📊 今日轉換公司債掃描完成\n"
        f"日期：{today:%Y-%m-%d}\n"
        f"總筆數：{n_rows} 檔\n"
        "\n"
        "📌 最後 20 筆詳細資料：\n"
        f"{detail_str}"
    )
    return msg



# ========= 整合主程式 =========

def main():
    today = dt.date.today()

    # 1) 先執行「程式1」：抓金管會 + 篩選轉換公司債 + 輸出 CSV
    csv_path = generate_convbond_csv_for_today()
    if csv_path is None:
        info_msg = f"📭 今日 ({today:%Y-%m-%d}) 無任何『轉換公司債』申報案件。"
        print(info_msg)
        send_line_message(info_msg)
        return

    # 2) 接著執行「程式2」：讀取 CSV，補股價 & 成交量，輸出 with_price / last20
    with_price_path, last20_path, n_rows = fill_prices_for_file(csv_path)

    # 3) 組訊息，推播到 LINE
    text = build_summary_message(today, csv_path, with_price_path, last20_path, n_rows)
    send_line_message(text)


if __name__ == "__main__":
    main()

